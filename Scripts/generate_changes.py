import json
import os
import re
import sys
from openpyxl import load_workbook

sys.stdout.reconfigure(encoding='utf-8')

def get_cell_val(sheet, cell_loc):
    val = sheet[cell_loc].value
    return str(val).strip() if val is not None else ""

def clean_name_title(rep_str):
    parts = re.split(r'[-;]', rep_str, maxsplit=1)
    if len(parts) == 2:
        return parts[0].strip(), parts[1].strip()
    return rep_str, ""

def main():
    excel_path = r"Templates/Mạng nội bộ UBND Xã Khâm Đức.xlsx"
    if not os.path.exists(excel_path):
        print(f"Error: Excel file not found at {excel_path}")
        return

    print(f"Loading Excel file: {excel_path}")
    wb = load_workbook(excel_path, data_only=True)
    
    # ----------------------------------------------------
    # 1. Parse Sheet "1. Tổng quan"
    # ----------------------------------------------------
    sheet_overview = wb["1. Tổng quan"]
    
    owner_org = get_cell_val(sheet_overview, "C6")
    owner_rep_raw = get_cell_val(sheet_overview, "C8")
    owner_addr = get_cell_val(sheet_overview, "C9")
    owner_phone = get_cell_val(sheet_overview, "C10")
    
    op_org = get_cell_val(sheet_overview, "C12")
    op_func = get_cell_val(sheet_overview, "C13")
    op_rep_raw = get_cell_val(sheet_overview, "C14")
    op_addr = get_cell_val(sheet_overview, "C15")
    op_phone = get_cell_val(sheet_overview, "C16")
    
    sys_purpose = get_cell_val(sheet_overview, "C18")
    sys_users = get_cell_val(sheet_overview, "C19")
    
    owner_rep_name, owner_rep_title = clean_name_title(owner_rep_raw)
    op_rep_name, op_rep_title = clean_name_title(op_rep_raw)

    print(f"  Owner Org: {owner_org}")
    print(f"  Operator Org: {op_org}")

    # ----------------------------------------------------
    # 2. Parse Sheet "3. Danh mục hệ thống"
    # ----------------------------------------------------
    sheet_sys = wb["3. Danh mục hệ thống"]
    
    # Extract devices (Rows 14 to 23)
    devices = []
    start_row = 14
    end_row = 23
    for r in range(start_row, end_row + 1):
        stt = sheet_sys.cell(row=r, column=1).value
        name = sheet_sys.cell(row=r, column=2).value
        zone = sheet_sys.cell(row=r, column=3).value
        qty = sheet_sys.cell(row=r, column=4).value
        purpose = sheet_sys.cell(row=r, column=5).value
        
        if name:
            devices.append({
                "stt": str(stt).strip() if stt else "",
                "name": str(name).strip(),
                "zone": str(zone).strip() if zone else "",
                "qty": str(qty).strip() if qty else "1",
                "purpose": str(purpose).strip() if purpose else ""
            })
            
    print(f"  Found {len(devices)} devices in survey.")

    # Find the Firewall model, Switch Core model from devices list
    fw_model = "Firewall Fortigate FG 100F"
    core_sw_model = "Switch L3 Maipu S3330-28TXF-AC"
    l2_sw_model = "Switch L2 Maipu (SL:08), Grandstream (SL:01)"
    ap_model = "Wireless AP Maipu (SL:30)"
    
    for dev in devices:
        name_lower = dev["name"].lower()
        if "firewall" in name_lower or "fortigate" in name_lower:
            fw_model = dev["name"]
        elif "switch l3" in name_lower or "s3330" in name_lower:
            core_sw_model = dev["name"]

    # Extract IP ranges (Rows 58 to 63, Private IP in col 5 (E), Public IP in col 6 (F))
    ip_ranges = []
    start_ip_row = 58
    end_ip_row = 63
    for r in range(start_ip_row, end_ip_row + 1):
        stt = sheet_sys.cell(row=r, column=1).value
        zone = sheet_sys.cell(row=r, column=2).value
        ip_private = sheet_sys.cell(row=r, column=5).value
        ip_public = sheet_sys.cell(row=r, column=6).value
        
        if zone:
            ip_ranges.append({
                "stt": str(stt).strip() if stt else "",
                "zone": str(zone).strip(),
                "ip_private": str(ip_private).strip() if ip_private else "",
                "ip_public": str(ip_public).strip() if ip_public else ""
            })
            
    print(f"  Found {len(ip_ranges)} IP network ranges.")

    # ----------------------------------------------------
    # 3. Parse Sheet "4. Khảo sát yêu cầu kỹ thuật"
    # ----------------------------------------------------
    sheet_survey = wb["4. Khảo sát yêu cầu kỹ thuật"]
    survey_answers = {}
    for r in range(6, 180):
        stt_val = sheet_survey.cell(row=r, column=1).value
        ans_val = sheet_survey.cell(row=r, column=3).value
        if stt_val is not None:
            try:
                stt_num = int(stt_val)
                survey_answers[stt_num] = str(ans_val).strip() if ans_val else ""
            except ValueError:
                pass

    # ----------------------------------------------------
    # 4. Generate changes.json
    # ----------------------------------------------------
    changes = {
        "target_file": "sessions/session_kham_duc/temp/filled_temp.docx",
        "show_ui": False,
        "replacements": []
    }
    
    # --- Paragraphs (General profile details) ---
    changes["replacements"].extend([
        {
            "comment": "Sub-header date year update",
            "scope": "paragraph",
            "anchor": "Đà Nẵng – 20xx",
            "find": "20xx",
            "replace": "2026"
        },
        {
            "comment": "Owner representative name and title",
            "scope": "paragraph",
            "anchor": "Người đại diện: Lê Ngọc Quang; chức vụ: Bí thư",
            "find": "Lê Ngọc Quang; chức vụ: Bí thư",
            "replace": f"{owner_rep_name}; chức vụ: {owner_rep_title}"
        },
        {
            "comment": "Owner address",
            "scope": "paragraph",
            "anchor": "Địa chỉ: số 72 Bạch Đằng, Phường Hải Châu, Thành phố Đà Nẵng.",
            "find": "số 72 Bạch Đằng, Phường Hải Châu, Thành phố Đà Nẵng.",
            "replace": owner_addr
        },
        {
            "comment": "Owner phone",
            "scope": "paragraph",
            "anchor": "Số điện thoại: (+84.236) 0805.1342",
            "find": "(+84.236) 0805.1342",
            "replace": owner_phone
        },
        {
            "comment": "Operator organization name",
            "scope": "paragraph",
            "anchor": "Tên tổ chức: UBND Phường Hải Châu",
            "find": "UBND Phường Hải Châu",
            "replace": op_org
        },
        {
            "comment": "Operator function regulation doc",
            "scope": "paragraph",
            "anchor": "Quy định chức năng, nhiệm vụ và quyền hạn: (Quyết định quy định chức năng, nhiệm vụ và quyền hạn)",
            "find": "(Quyết định quy định chức năng, nhiệm vụ và quyền hạn)",
            "replace": op_func
        },
        {
            "comment": "Operator representative",
            "scope": "paragraph",
            "anchor": "Người đại diện: (Tên Người đại diện), Chức vụ: (Chức vụ)",
            "find": "(Tên Người đại diện), Chức vụ: (Chức vụ)",
            "replace": f"{op_rep_name}, Chức vụ: {op_rep_title}"
        },
        {
            "comment": "Operator address",
            "scope": "paragraph",
            "anchor": "Địa chỉ: số 15 Lê Hồng Phong, Phường Hải Châu, Thành phố Đà Nẵng",
            "find": "số 15 Lê Hồng Phong, Phường Hải Châu, Thành phố Đà Nẵng",
            "replace": op_addr
        },
        {
            "comment": "Operator phone",
            "scope": "paragraph",
            "anchor": "Số điện thoại: (Số điện thoại)",
            "find": "(Số điện thoại)",
            "replace": op_phone
        },
        {
            "comment": "Scope location",
            "scope": "paragraph",
            "anchor": "Phạm vi, quy mô của Hệ thống thông tin: triển khai tại trụ sở phục vụ hoạt động nội bộ của cơ quan.",
            "find": "triển khai tại trụ sở phục vụ hoạt động nội bộ của cơ quan.",
            "replace": f"triển khai tại trụ sở {op_org}, phục vụ hoạt động nội bộ của cơ quan."
        },
        {
            "comment": "Users scope",
            "scope": "paragraph",
            "anchor": "Đối tượng phục vụ của hệ thống: Công chức, viên chức, người lao động tại UBND Phường Hải Châu.",
            "find": "UBND Phường Hải Châu",
            "replace": op_org
        },
        {
            "comment": "Policy organization in Phụ lục I",
            "scope": "paragraph",
            "anchor": "Ban hành Quyết định số PC ngày … của (Tên cơ quan) về việc phân công nhiệm vụ bảo vệ an ninh mạng",
            "find": "(Tên cơ quan)",
            "replace": op_org
        }
    ])

    # --- Cover Page in Table 0 cell(0,0) ---
    changes["replacements"].extend([
        {
            "comment": "Cover: Operator Org",
            "scope": "table_cell",
            "table_index": 0,
            "row_index": 0,
            "col_index": 0,
            "find": "UBND PHƯỜNG HẢI CHÂU",
            "replace": op_org.upper()
        },
        {
            "comment": "Cover: Year",
            "scope": "table_cell",
            "table_index": 0,
            "row_index": 0,
            "col_index": 0,
            "find": "20xx",
            "replace": "2026"
        }
    ])

    # --- Table 2: Bảng 1. Danh mục các HTTT thành phần ---
    changes["replacements"].extend([
        {
            "comment": "Table 1: STT",
            "scope": "table_cell",
            "table_index": 2,
            "row_index": 1,
            "col_index": 0,
            "find": "",
            "replace": "1"
        }
    ])

    # --- Table 3: Bảng 2. Danh mục thiết bị ---
    # We will write all devices extracted. The new word_editor will automatically add rows!
    for idx, dev in enumerate(devices):
        r_idx = idx + 1 # rows start after title row 0
        dev_desc = f"{dev['name']} (SL: {dev['qty'].zfill(2)})"
        
        # In the template, rows 1, 2, 3, 4 have placeholders. For these rows, we specify 'find'
        # so that it replaces the placeholder. For subsequent rows, 'find' is empty (direct write).
        find_val = ""
        if r_idx == 1:
            find_val = "FirewallCore/ Hãng sản xuất model"
        elif r_idx == 2:
            find_val = "SWCore/ Hãng sản xuất model"
        elif r_idx == 3:
            find_val = "SWL2/ Hãng sản xuất model (SL:03)"
        elif r_idx == 4:
            find_val = "AP/ Hãng sản xuất model (SL:03)"
            
        changes["replacements"].extend([
            {
                "comment": f"Table 2: Row {r_idx} STT",
                "scope": "table_cell",
                "table_index": 3,
                "row_index": r_idx,
                "col_index": 0,
                "find": "", # direct write
                "replace": str(idx + 1)
            },
            {
                "comment": f"Table 2: Row {r_idx} Device Name",
                "scope": "table_cell",
                "table_index": 3,
                "row_index": r_idx,
                "col_index": 1,
                "find": find_val,
                "replace": dev_desc
            },
            {
                "comment": f"Table 2: Row {r_idx} Zone",
                "scope": "table_cell",
                "table_index": 3,
                "row_index": r_idx,
                "col_index": 2,
                "find": "Vùng mạng biên" if r_idx == 1 else ("Vùng mạng nội bộ" if r_idx in (2,3,4) else ""),
                "replace": dev["zone"]
            },
            {
                "comment": f"Table 2: Row {r_idx} Purpose",
                "scope": "table_cell",
                "table_index": 3,
                "row_index": r_idx,
                "col_index": 3,
                "find": "Quản lý truy cập vào/ra" if r_idx == 1 else ("Thiết bị định tuyến quản lý, kết nối vùng mạng nội bộ" if r_idx == 2 else ("Thiết bị chuyển mạch" if r_idx == 3 else ("Thiết bị phát wifi" if r_idx == 4 else ""))),
                "replace": dev["purpose"]
            }
        ])

    # --- Table 6: Bảng 5. Quy hoạch địa chỉ IP ---
    for idx, ip in enumerate(ip_ranges):
        r_idx = idx + 1
        find_zone = ""
        find_priv = ""
        find_pub = ""
        
        if r_idx == 1:
            find_zone = "Vùng mạng nội bộ"
            find_priv = "192.168.1.0/24"
        elif r_idx == 2:
            find_zone = "Vùng mạng không dây cho CBBCCVC"
            find_priv = "192.168.2.0/24"
        elif r_idx == 3:
            find_zone = "Vùng mạng không dây cho khách"
            find_priv = "192.168.3.0/24"
        elif r_idx == 4:
            find_zone = "Vùng mạng biên"
            find_pub = "203.119.x.x"
            
        changes["replacements"].extend([
            {
                "comment": f"Table 5: Row {r_idx} STT",
                "scope": "table_cell",
                "table_index": 6,
                "row_index": r_idx,
                "col_index": 0,
                "find": "",
                "replace": str(idx + 1)
            },
            {
                "comment": f"Table 5: Row {r_idx} Zone Name",
                "scope": "table_cell",
                "table_index": 6,
                "row_index": r_idx,
                "col_index": 1,
                "find": find_zone,
                "replace": ip["zone"]
            },
            {
                "comment": f"Table 5: Row {r_idx} IP Private",
                "scope": "table_cell",
                "table_index": 6,
                "row_index": r_idx,
                "col_index": 2,
                "find": find_priv,
                "replace": ip["ip_private"]
            },
            {
                "comment": f"Table 5: Row {r_idx} IP Public",
                "scope": "table_cell",
                "table_index": 6,
                "row_index": r_idx,
                "col_index": 3,
                "find": find_pub,
                "replace": ip["ip_public"]
            }
        ])

    # --- Table 7: Bảng 6. Danh mục hệ thống và cấp độ ---
    changes["replacements"].extend([
        {
            "comment": "Table 6: Row 1 STT",
            "scope": "table_cell",
            "table_index": 7,
            "row_index": 1,
            "col_index": 0,
            "find": "",
            "replace": "1"
        }
    ])

    # --- Phụ lục I Table 9 (Policy organization) ---
    changes["replacements"].extend([
        {
            "comment": "Phụ lục I: Policy organization",
            "scope": "table_cell",
            "table_index": 9,
            "row_index": 2,
            "col_index": 1,
            "find": "(Tên cơ quan)",
            "replace": op_org
        }
    ])

    # --- Phụ lục II Table 52 (Technical protection overview) ---
    ans_1 = survey_answers.get(1, f"Sử dụng {fw_model} có tích hợp chức năng VPN để quản lý truy cập, quản trị hệ thống từ xa an toàn.")
    ans_2 = survey_answers.get(6, f"Cấu hình ACL và sử dụng tính năng IPS/IDS trên {fw_model} để quản lý truy cập giữa các vùng mạng và phòng chống xâm nhập.")
    ans_3 = survey_answers.get(3, "Sử dụng Kaspersky Internet Security")
    ans_5 = survey_answers.get(4, "Sử dụng hệ thống thư điện tử tập trung do Thành phố cung cấp")
    ans_6 = survey_answers.get(2, f"Có phương án dự phòng cho thiết bị tường lửa {fw_model}; thiết bị chuyển mạch Core Switch chưa có phương án dự phòng.")

    changes["replacements"].extend([
        {
            "comment": "Phụ lục II.1.b: VPN solution",
            "scope": "table_cell",
            "table_index": 52,
            "row_index": 1,
            "col_index": 3,
            "find": "FirewallCore",
            "replace": fw_model
        },
        {
            "comment": "Phụ lục II.1.b: ACL & IPS",
            "scope": "table_cell",
            "table_index": 52,
            "row_index": 2,
            "col_index": 3,
            "find": "Cấu hình ACL trên FirewallCore để quản lý truy cập giữa các vùng mạng và phòng chống xâm nhập.",
            "replace": ans_2
        },
        {
            "comment": "Phụ lục II.1.b: AV software",
            "scope": "table_cell",
            "table_index": 52,
            "row_index": 3,
            "col_index": 3,
            "find": "(Tên phần mềm Antivirus)",
            "replace": ans_3
        },
        {
            "comment": "Phụ lục II.1.b: Email protection",
            "scope": "table_cell",
            "table_index": 52,
            "row_index": 5,
            "col_index": 3,
            "find": "Không có hệ thống thư điện tử",
            "replace": ans_5
        },
        {
            "comment": "Phụ lục II.1.b: Main device backup",
            "scope": "table_cell",
            "table_index": 52,
            "row_index": 6,
            "col_index": 3,
            "find": "Sử dụng các thiết bị phổ biến trên thị trường",
            "replace": ans_6
        }
    ])

    # --- Phụ lục II Table 53 (External access control) ---
    changes["replacements"].extend([
        {
            "comment": "Phụ lục II.1.2: VPN device replacement",
            "scope": "table_cell",
            "table_index": 53,
            "row_index": 1,
            "col_index": 3,
            "find": "FirewallCore",
            "replace": fw_model
        },
        {
            "comment": "Phụ lục II.1.2: ACL device replacement",
            "scope": "table_cell",
            "table_index": 53,
            "row_index": 2,
            "col_index": 3,
            "find": "FirewallCore",
            "replace": fw_model
        },
        {
            "comment": "Phụ lục II.1.2: Timeout device replacement",
            "scope": "table_cell",
            "table_index": 53,
            "row_index": 3,
            "col_index": 3,
            "find": "FirewallCore",
            "replace": fw_model
        }
    ])

    # --- Phụ lục II Table 54 (Internal access control) ---
    changes["replacements"].extend([
        {
            "comment": "Phụ lục II.1.3: Access control device",
            "scope": "table_cell",
            "table_index": 54,
            "row_index": 1,
            "col_index": 3,
            "find": "FirewallCore",
            "replace": fw_model
        }
    ])

    # --- Phụ lục II Table 55 (System logging devices) ---
    changes["replacements"].extend([
        {
            "comment": "Phụ lục II.1.4: Firewall logging",
            "scope": "table_cell",
            "table_index": 55,
            "row_index": 2,
            "col_index": 0,
            "find": "FirewallCore",
            "replace": fw_model
        },
        {
            "comment": "Phụ lục II.1.4: Core Switch logging",
            "scope": "table_cell",
            "table_index": 55,
            "row_index": 3,
            "col_index": 0,
            "find": "SWCore",
            "replace": core_sw_model
        },
        {
            "comment": "Phụ lục II.1.4: L2 Switch logging",
            "scope": "table_cell",
            "table_index": 55,
            "row_index": 4,
            "col_index": 0,
            "find": "SWL2 (SL:03)",
            "replace": l2_sw_model
        },
        {
            "comment": "Phụ lục II.1.4: AP logging",
            "scope": "table_cell",
            "table_index": 55,
            "row_index": 5,
            "col_index": 0,
            "find": "AP (SL:03)",
            "replace": ap_model
        }
    ])

    # --- Phụ lục II Table 56 (IPS protection) ---
    changes["replacements"].extend([
        {
            "comment": "Phụ lục II.1.5: IPS zone defense",
            "scope": "table_cell",
            "table_index": 56,
            "row_index": 1,
            "col_index": 3,
            "find": "FirewallCore",
            "replace": fw_model
        },
        {
            "comment": "Phụ lục II.1.5: IPS DB update",
            "scope": "table_cell",
            "table_index": 56,
            "row_index": 2,
            "col_index": 3,
            "find": "FirewallCore",
            "replace": fw_model
        }
    ])

    # --- Phụ lục II Table 57 (Device protection) ---
    changes["replacements"].extend([
        {
            "comment": "Phụ lục II.1.6: Firewall protection",
            "scope": "table_cell",
            "table_index": 57,
            "row_index": 2,
            "col_index": 0,
            "find": "FirewallCore",
            "replace": fw_model
        },
        {
            "comment": "Phụ lục II.1.6: Core Switch protection",
            "scope": "table_cell",
            "table_index": 57,
            "row_index": 3,
            "col_index": 0,
            "find": "SWCore",
            "replace": core_sw_model
        },
        {
            "comment": "Phụ lục II.1.6: L2 Switch protection",
            "scope": "table_cell",
            "table_index": 57,
            "row_index": 4,
            "col_index": 0,
            "find": "SWL2 (SL:03)",
            "replace": l2_sw_model
        },
        {
            "comment": "Phụ lục II.1.6: AP protection",
            "scope": "table_cell",
            "table_index": 57,
            "row_index": 5,
            "col_index": 0,
            "find": "AP (SL:03)",
            "replace": ap_model
        }
    ])

    # ----------------------------------------------------
    # Save changes.json
    # ----------------------------------------------------
    output_dir = "sessions/session_kham_duc"
    os.makedirs(output_dir, exist_ok=True)
    output_path = os.path.join(output_dir, "changes.json")
    
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(changes, f, ensure_ascii=False, indent=2)
        
    print(f"[SUCCESS] Generated dynamic changes.json at {output_path}")

if __name__ == "__main__":
    main()
