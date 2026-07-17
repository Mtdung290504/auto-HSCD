import sys
import os
from datetime import datetime, date, time
import openpyxl

# Set UTF-8 encoding for console output on Windows
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")
if hasattr(sys.stderr, "reconfigure"):
    sys.stderr.reconfigure(encoding="utf-8")


def clean_cell_value(val):
    if val is None:
        return ""

    # Format dates/times so the model sees a plain readable string,
    # not "2026-07-17 00:00:00" from str(datetime).
    if isinstance(val, datetime):
        if val.time() == time(0, 0):
            val_str = val.strftime("%Y-%m-%d")
        else:
            val_str = val.strftime("%Y-%m-%d %H:%M:%S")
    elif isinstance(val, date):
        val_str = val.strftime("%Y-%m-%d")
    elif isinstance(val, time):
        val_str = val.strftime("%H:%M:%S")
    elif isinstance(val, float) and val.is_integer():
        # Avoid "1.0" for what is really the integer 1
        val_str = str(int(val))
    else:
        val_str = str(val)

    # Escape/strip characters that would break a markdown table row
    val_str = (
        val_str.replace("|", "\\|")
        .replace("\r\n", " ")
        .replace("\n", " ")
        .replace("\r", " ")
    )
    return val_str.strip()


def build_merge_value_map(sheet):
    """Merged cells only carry a value in their top-left cell; openpyxl
    returns None for the rest (and disallows writing to them directly).
    Build a lookup so every cell in a merge range resolves to that value,
    except child cells which resolve to a coordinate list [row, col]."""
    merge_map = {}
    for merged_range in sheet.merged_cells.ranges:
        min_r, min_c = merged_range.min_row, merged_range.min_col
        max_r, max_c = merged_range.max_row, merged_range.max_col

        top_left_val = sheet.cell(min_r, min_c).value
        merge_map[(min_r, min_c)] = top_left_val

        # Child cells in the merge block point back to the master cell via [row, col]
        for r in range(min_r, max_r + 1):
            for c in range(min_c, max_c + 1):
                if r == min_r and c == min_c:
                    continue
                merge_map[(r, c)] = [min_r - 1, min_c - 1]
    return merge_map


def sheet_to_grid(sheet):
    merge_map = build_merge_value_map(sheet)

    grid = []
    for r in range(1, sheet.max_row + 1):
        row_data = []
        for c in range(1, sheet.max_column + 1):
            val = merge_map.get((r, c), sheet.cell(row=r, column=c).value)
            if isinstance(val, list):
                row_data.append(val)
            else:
                row_data.append(clean_cell_value(val))
        grid.append(row_data)

    # Strip trailing empty rows
    while grid and all(v == "" for v in grid[-1]):
        grid.pop()
    if not grid:
        return []

    # Pad every row to the same width BEFORE filtering columns, otherwise
    # column filtering misaligns rows of different lengths.
    max_cols = max(len(row) for row in grid)
    grid = [row + [""] * (max_cols - len(row)) for row in grid]

    # Find columns that are empty across all rows, and drop them
    col_is_empty = [all(row[c] == "" for row in grid) for c in range(max_cols)]
    clean_grid = [
        [row[c] for c in range(max_cols) if not col_is_empty[c]] for row in grid
    ]

    return clean_grid


def render_sheet_markdown(sheet):
    lines = []
    status = " (Hidden)" if sheet.sheet_state != "visible" else ""
    lines.append(f"## Sheet: {sheet.title}{status}")

    if sheet.max_row == 0 or sheet.max_column == 0:
        lines.append("*Sheet is empty.*\n")
        return "\n".join(lines)

    grid = sheet_to_grid(sheet)
    if not grid or not grid[0]:
        lines.append("*Sheet has only empty cells.*\n")
        return "\n".join(lines)

    def get_real_val(r_idx, c_idx):
        cell_val = grid[r_idx][c_idx]
        if isinstance(cell_val, list):
            ref_r, ref_c = cell_val
            return get_real_val(ref_r, ref_c)
        return cell_val

    headers = [get_real_val(0, i) for i in range(len(grid[0]))]
    headers = [h if h else f"Col {i+1}" for i, h in enumerate(headers)]
    lines.append("| " + " | ".join(headers) + " |")
    lines.append("| " + " | ".join(["---"] * len(headers)) + " |")

    for r_idx, row in enumerate(grid[1:]):
        row_vals = [get_real_val(r_idx + 1, c_idx) for c_idx in range(len(row))]
        lines.append("| " + " | ".join(row_vals) + " |")
    lines.append("")

    return "\n".join(lines)


def excel_to_markdown(filepath, include_hidden=False):
    if not os.path.exists(filepath):
        print(f"Error: File not found: {filepath}")
        sys.exit(1)

    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
    except Exception as e:
        print(f"Error loading Excel file: {e}")
        sys.exit(1)

    print(f"# EXCEL DATA DUMP: {os.path.basename(filepath)}\n")

    any_sheet_shown = False
    for sheet in wb.worksheets:
        is_hidden = sheet.sheet_state != "visible"
        if is_hidden and not include_hidden:
            continue
        any_sheet_shown = True
        try:
            print(render_sheet_markdown(sheet))
        except Exception as e:
            print(f"## Sheet: {sheet.title}\n*Error reading this sheet: {e}*\n")

    if not any_sheet_shown:
        print("*No visible sheets found. Re-run with --all to include hidden sheets.*")


def excel_to_json(filepath, output_path, include_hidden=False):
    if not os.path.exists(filepath):
        print(f"Error: File not found: {filepath}")
        sys.exit(1)

    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
    except Exception as e:
        print(f"Error loading Excel file: {e}")
        sys.exit(1)

    result = {"sheets": {}}

    for sheet in wb.worksheets:
        is_hidden = sheet.sheet_state != "visible"
        if is_hidden and not include_hidden:
            continue
        try:
            grid = sheet_to_grid(sheet)
            result["sheets"][sheet.title] = {"hidden": is_hidden, "grid": grid}
        except Exception as e:
            print(f"Error reading sheet {sheet.title}: {e}")

    import json

    try:
        # Resolve output directory
        out_dir = os.path.dirname(output_path)
        if out_dir:
            os.makedirs(out_dir, exist_ok=True)
        with open(output_path, "w", encoding="utf-8") as f:
            json.dump(result, f, ensure_ascii=False, separators=(",", ":"))
        print(f"[SUCCESS] Saved Excel JSON data to: {output_path}")
    except Exception as e:
        print(f"Error writing JSON to file: {e}")
        sys.exit(1)


if __name__ == "__main__":
    args = sys.argv[1:]
    if not args or args[0] in ("-h", "--help"):
        print("Usage:")
        print("  Mặc định (Xuất JSON ra file):")
        print(
            "    python excel_parser.py <path_to_excel_file> <output_json_file> [--all]"
        )
        print("  Xuất Markdown (ra stdout):")
        print("    python excel_parser.py <path_to_excel_file> --md [--all]")
        sys.exit(0)

    excel_path = args[0]
    include_hidden = "--all" in args

    if "--md" in args or "--markdown" in args:
        excel_to_markdown(excel_path, include_hidden)
    else:
        remaining_args = [a for a in args[1:] if a != "--all"]
        if not remaining_args:
            print("Error: Default mode requires an output JSON file path.")
            print(
                "Usage: python excel_parser.py <path_to_excel_file> <output_json_file> [--all]"
            )
            sys.exit(1)
        output_json_path = remaining_args[0]
        excel_to_json(excel_path, output_json_path, include_hidden)
